# Import NumPy for numerical operations
# Importa NumPy para operações numéricas
import numpy as np

# Import Excel workbook creation tools from openpyxl
# Importa ferramentas para criação de planilhas Excel do openpyxl
from openpyxl import Workbook

# Import font utilities to format Excel cells (e.g., bold text)
# Importa utilidades de fonte para formatar células do Excel (ex: texto em negrito)
from openpyxl.styles import Font


class MassPostProcessor:
    """
    Post-processing utilities for the mass balance model.
    Supports ANY number of components.
    """
    """
    Utilidades de pós-processamento para o modelo de balanço de massa.
    Suporta QUALQUER número de componentes.
    """

    def export_results_to_excel(
        self,
        filename,
        case_name,
        components,
        FRet,
        ZRet,
        FPerm,
        ZPerm,
        FMemb,
        FMemb_Comp,
        ZMemb,
        PRetCell,
        PPermCell
    ):

        # Number of components in the mixture
        # Número de componentes na mistura
        n_comp = len(components)

        # Number of axial nodes (segments + inlet)
        # Número de nós axiais (segmentos + entrada)
        N = len(FRet) - 1

        # Create Excel workbook
        # Cria um arquivo Excel
        wb = Workbook()

        # Select active worksheet
        # Seleciona a planilha ativa
        ws = wb.active

        # Rename worksheet
        # Renomeia a planilha
        ws.title = "MassModel"

        # ------------------------------
        # Case info
        # Informações do caso
        # ------------------------------

        # Write case label
        # Escreve o rótulo do caso
        ws["A1"] = "Case:"

        # Write case name
        # Escreve o nome do caso
        ws["B1"] = case_name

        # Make case name bold
        # Deixa o nome do caso em negrito
        ws["B1"].font = Font(bold=True)

        # Write component label
        # Escreve rótulo de componentes
        ws["A2"] = "Components:"

        # Join component names as a single string
        # Junta os nomes dos componentes em uma única string
        ws["B2"] = ", ".join(components)

        # Insert empty row for spacing
        # Insere uma linha vazia para espaçamento
        ws.append([])

        # ------------------------------
        # Header
        # Cabeçalho da tabela
        # ------------------------------

        # Initialize header with axial node index and total retentate flow
        # Inicializa o cabeçalho com índice axial e vazão total do retentado
        header = ["k", "F_tot [mol/s]"]

        # Add retentate component flows
        # Adiciona vazões molares por componente no retentado
        for i in range(n_comp):
            header.append(f"FRet_{i}")

        # Add retentate compositions
        # Adiciona composições molares do retentado
        for i in range(n_comp):
            header.append(f"ZRet[{i}]")

        # Add retentate pressure column
        # Adiciona coluna de pressão do retentado
        header.append("PRet [Pa]")

        # Add total permeate flow
        # Adiciona vazão total do permeado
        header.append("FPerm_tot [mol/s]")

        # Add permeate component flows
        # Adiciona vazões molares por componente no permeado
        for i in range(n_comp):
            header.append(f"FPerm_{i}")

        # Add permeate compositions
        # Adiciona composições molares do permeado
        for i in range(n_comp):
            header.append(f"ZPerm[{i}]")

        # Add permeate pressure column
        # Adiciona coluna de pressão do permeado
        header.append("PPerm [Pa]")

        # Add total membrane flux
        # Adiciona fluxo total através da membrana
        header.append("FMemb_tot [mol/s]")

        # Add membrane flux per component
        # Adiciona fluxo da membrana por componente
        for i in range(n_comp):
            header.append(f"FMemb_{i}")

        # Add membrane composition
        # Adiciona composição do fluxo através da membrana
        for i in range(n_comp):
            header.append(f"ZMemb[{i}]")

        # Write header row to Excel
        # Escreve o cabeçalho na planilha Excel
        ws.append(header)

        # ------------------------------
        # Rows
        # Linhas da tabela
        # ------------------------------

        # Loop over axial nodes
        # Loop sobre os nós axiais
        for k in range(N + 1):

            # Initialize row with node index and total retentate flow
            # Inicializa a linha com índice do nó e vazão total do retentado
            row = [k, float(FRet[k])]

            # ------------------------------
            # Retentate component flows
            # Vazões por componente no retentado
            # ------------------------------

            for i in range(n_comp):

                # Compute component flow = total flow × composition
                # Calcula vazão do componente = vazão total × composição
                row.append(float(FRet[k] * ZRet[k, i]))

            # ------------------------------
            # Retentate compositions
            # Composições do retentado
            # ------------------------------

            for i in range(n_comp):
                row.append(float(ZRet[k, i]))

            # Add retentate pressure
            # Adiciona pressão do retentado
            row.append(float(PRetCell[k]))

            # Add total permeate flow
            # Adiciona vazão total do permeado
            row.append(float(FPerm[k]))

            # ------------------------------
            # Permeate component flows
            # Vazões por componente no permeado
            # ------------------------------

            for i in range(n_comp):

                # Component permeate flow = total permeate × composition
                # Vazão do componente no permeado = vazão total × composição
                row.append(float(FPerm[k] * ZPerm[k, i]))

            # ------------------------------
            # Permeate compositions
            # Composições do permeado
            # ------------------------------

            for i in range(n_comp):
                row.append(float(ZPerm[k, i]))

            # Add permeate pressure
            # Adiciona pressão do permeado
            row.append(float(PPermCell[k]))

            # Add total membrane flux
            # Adiciona fluxo total através da membrana
            row.append(float(FMemb[k]))

            # ------------------------------
            # Membrane flux per component
            # Fluxo da membrana por componente
            # ------------------------------

            for i in range(n_comp):
                row.append(float(FMemb_Comp[k, i]))

            # ------------------------------
            # Membrane composition
            # Composição do fluxo da membrana
            # ------------------------------

            for i in range(n_comp):
                row.append(float(ZMemb[k, i]))

            # Write row to Excel
            # Escreve a linha na planilha Excel
            ws.append(row)

        # Save Excel file
        # Salva o arquivo Excel
        wb.save(filename)
