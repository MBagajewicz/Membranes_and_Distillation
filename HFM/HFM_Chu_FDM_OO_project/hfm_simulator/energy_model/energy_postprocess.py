# Import NumPy for numerical operations
# Importa NumPy para operações numéricas
import numpy as np

# Import CoolProp function for thermodynamic property calculations
# Importa função do CoolProp para cálculos de propriedades termodinâmicas
from CoolProp.CoolProp import PropsSI

# Import Excel workbook creation tools
# Importa ferramentas para criação de planilhas Excel
from openpyxl import Workbook

# Import Excel font utilities for formatting
# Importa utilidades de fonte do Excel para formatação
from openpyxl.styles import Font


class EnergyPostProcessor:
    """
    Utility class responsible for post-processing energy model results.

    This class computes additional thermodynamic quantities and exports
    simulation results to Excel.

    Classe utilitária responsável pelo pós-processamento dos resultados
    do modelo de energia.

    Esta classe calcula propriedades termodinâmicas adicionais e exporta
    resultados da simulação para Excel.
    """

    def __init__(self, components):

        # List of components in the mixture
        # Lista de componentes da mistura
        self.components = components

        # Number of components
        # Número de componentes
        self.n_comp = len(components)

    # -------------------------------------------------
    # Dew point for multicomponent mixture
    # Ponto de orvalho para mistura multicomponente
    # -------------------------------------------------

    def dew_point(self, P, z):

        try:

            # Build CoolProp mixture string
            # Constrói string de mistura para o CoolProp
            mixture = "HEOS::"

            for i, comp in enumerate(self.components):

                # Add separator between components
                # Adiciona separador entre componentes
                if i > 0:
                    mixture += "&"

                # Append component with mole fraction
                # Adiciona componente com fração molar
                mixture += f"{comp}[{z[i]}]"

            # Compute dew point temperature at pressure P
            # Calcula temperatura de ponto de orvalho na pressão P
            Tdew = PropsSI(
                "T",
                "P", P,
                "Q", 1,
                mixture
            )

            return Tdew

        except:

            # Return NaN if CoolProp fails
            # Retorna NaN se o CoolProp falhar
            return np.nan

    # -------------------------------------------------
    # Excel export
    # Exportação para Excel
    # -------------------------------------------------

    def export_results_to_excel(
        self,
        filename,
        case_name,
        components,
        FRet, ZRet, hRet, hRet_mix,
        FPerm, ZPerm, hPerm, hPerm_mix,
        FMemb, FMemb_comp, ZMemb, hMemb,
        PRetCell, PPermCell, T_ret, T_per,
        UA
    ):

        # Number of components
        # Número de componentes
        n_comp = self.n_comp

        # Number of axial nodes
        # Número de nós axiais
        N = len(FRet) - 1

        # Create Excel workbook
        # Cria arquivo Excel
        wb = Workbook()

        # Get active worksheet
        # Obtém planilha ativa
        ws = wb.active

        # Rename worksheet
        # Renomeia a planilha
        ws.title = "EnergyModel"

        # Write case information
        # Escreve informações do caso
        ws["A1"] = "Case:"
        ws["B1"] = case_name
        ws["B1"].font = Font(bold=True)

        # Write component list
        # Escreve lista de componentes
        ws["A2"] = "Components:"
        ws["B2"] = ", ".join(components)

        # Insert empty row
        # Insere linha vazia
        ws.append([])

        # -------------------------------------------------
        # Header
        # Cabeçalho
        # -------------------------------------------------

        # Initialize header with axial index and total retentate flow
        # Inicializa cabeçalho com índice axial e vazão total do retentado
        header = ["k", "F_tot"]

        # Add retentate component flows
        # Adiciona vazões molares por componente no retentado
        for i in range(n_comp):
            header.append(f"FRet_{i}")

        # Add retentate compositions
        # Adiciona composições do retentado
        for i in range(n_comp):
            header.append(f"ZRet[{i}]")

        # Add retentate pressure and enthalpy
        # Adiciona pressão e entalpia do retentado
        header.append("PRet")
        header.append("hRet")

        # Add total permeate flow
        # Adiciona vazão total do permeado
        header.append("FPerm_tot")

        # Add permeate component flows
        # Adiciona vazões molares por componente no permeado
        for i in range(n_comp):
            header.append(f"FPerm_{i}")

        # Add permeate compositions
        # Adiciona composições do permeado
        for i in range(n_comp):
            header.append(f"ZPerm[{i}]")

        # Add permeate pressure and enthalpy
        # Adiciona pressão e entalpia do permeado
        header.append("PPerm")
        header.append("hPerm")

        # Add membrane total flux
        # Adiciona fluxo total da membrana
        header.append("FMemb_tot")

        # Add membrane component flux
        # Adiciona fluxo por componente da membrana
        for i in range(n_comp):
            header.append(f"FMemb_{i}")

        # Add membrane composition
        # Adiciona composição do fluxo da membrana
        for i in range(n_comp):
            header.append(f"ZMemb[{i}]")

        # Additional thermodynamic variables
        # Variáveis termodinâmicas adicionais
        header += [
            "T_ret",
            "T_per",
            "Q_cond",
            "Tdew_ret",
            "Tdew_per",
            "Tdew_mem"
        ]

        # Write header row
        # Escreve linha de cabeçalho
        ws.append(header)

        # -------------------------------------------------
        # Rows
        # Linhas de dados
        # -------------------------------------------------

        for k in range(N + 1):

            # Start row with axial node index and total retentate flow
            # Inicia linha com índice axial e vazão total do retentado
            row = [k, float(FRet[k])]

            # Retentate component flows
            # Vazões por componente no retentado
            for i in range(n_comp):
                row.append(float(FRet[k] * ZRet[k, i]))

            # Retentate compositions
            # Composições do retentado
            for i in range(n_comp):
                row.append(float(ZRet[k, i]))

            # Retentate pressure and enthalpy
            # Pressão e entalpia do retentado
            row.append(float(PRetCell[k]))
            row.append(float(hRet[k]))

            # Total permeate flow
            # Vazão total do permeado
            row.append(float(FPerm[k]))

            # Permeate component flows
            # Vazões por componente no permeado
            for i in range(n_comp):
                row.append(float(FPerm[k] * ZPerm[k, i]))

            # Permeate compositions
            # Composições do permeado
            for i in range(n_comp):
                row.append(float(ZPerm[k, i]))

            # Permeate pressure and enthalpy
            # Pressão e entalpia do permeado
            row.append(float(PPermCell[k]))
            row.append(float(hPerm[k]))

            # Total membrane flux
            # Fluxo total da membrana
            row.append(float(FMemb[k]))

            # Membrane component flux
            # Fluxo por componente na membrana
            for i in range(n_comp):
                row.append(float(FMemb_comp[k, i]))

            # Membrane composition
            # Composição do fluxo da membrana
            for i in range(n_comp):
                row.append(float(ZMemb[k, i]))

            # Heat conduction between streams
            # Condução de calor entre correntes
            if k > 0:
                Q_cond = UA[k] * (T_ret[k] - T_per[k-1])
            else:
                Q_cond = ""

            # Compute dew point temperatures
            # Calcula temperaturas de ponto de orvalho
            Tdew_ret = self.dew_point(PRetCell[k], ZRet[k])
            Tdew_per = self.dew_point(PPermCell[k], ZPerm[k])
            Tdew_mem = self.dew_point(PRetCell[k], ZMemb[k])

            # Append thermodynamic variables
            # Adiciona variáveis termodinâmicas
            row += [
                float(T_ret[k]),
                float(T_per[k]) if k < N else "",
                Q_cond,
                Tdew_ret,
                Tdew_per,
                Tdew_mem
            ]

            # Write row to Excel
            # Escreve linha na planilha
            ws.append(row)

        # Save Excel file
        # Salva arquivo Excel
        wb.save(filename)